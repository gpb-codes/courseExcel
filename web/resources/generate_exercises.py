import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__))

def save(wb, name):
    path = os.path.join(OUT, name)
    wb.save(path)
    print(f"  -> {name}")

def style_header(ws, row=1, cols=5):
    hf = Font(bold=True, color="FFFFFF", size=11)
    hf2 = PatternFill(start_color="6B21A8", end_color="6B21A8", fill_type="solid")
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hf
        cell.fill = hf2
        cell.alignment = Alignment(horizontal="center")

def auto_width(ws):
    for col in ws.columns:
        mx = 0
        for cell in col:
            try: mx = max(mx, len(str(cell.value or "")))
            except: pass
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx + 3

# --- M1: Primeros Pasos ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Datos Basicos"
ws["A1"] = "Nombre"; ws["B1"] = "Edad"; ws["C1"] = "Ciudad"
ws["A2"] = "Ana"; ws["B2"] = 28; ws["C2"] = "Madrid"
ws["A3"] = "Luis"; ws["B3"] = 35; ws["C3"] = "Barcelona"
ws["A4"] = "Sofia"; ws["B4"] = 22; ws["C4"] = "Valencia"
style_header(ws, 1, 3)
auto_width(ws)
save(wb, "M01_Primeros_Pasos.xlsx")

# --- M2: Manejo de Datos ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Series y Rangos"
ws["A1"] = "Mes"; ws["B1"] = "Ventas"
data = [("Enero", 12000), ("Febrero", 15000), ("Marzo", 11000), ("Abril", 16000), ("Mayo", 14000), ("Junio", 18000)]
for i, (m, v) in enumerate(data, 2):
    ws.cell(row=i, column=1, value=m)
    ws.cell(row=i, column=2, value=v)
style_header(ws, 1, 2)
auto_width(ws)
save(wb, "M02_Manejo_Datos.xlsx")

# --- M3: Formato y Presentacion ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Formato"
headers = ["Producto", "Precio", "Cantidad", "Total"]
ws.append(headers)
data = [("Laptop", 1200, 5), ("Mouse", 25, 20), ("Teclado", 80, 10), ("Monitor", 350, 4)]
for p, pr, c in data:
    ws.append([p, pr, c, pr*c])
style_header(ws, 1, 4)
moneda = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
for r in range(2, 6):
    ws.cell(row=r, column=2).fill = moneda
    ws.cell(row=r, column=4).fill = moneda
auto_width(ws)
save(wb, "M03_Formato_Presentacion.xlsx")

# --- M4: Formulas y Funciones ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Funciones"
ws.append(["Producto", "Enero", "Febrero", "Marzo", "Total"])
data = [("Laptop", 50, 65, 55), ("Mouse", 120, 110, 130), ("Teclado", 80, 85, 75), ("Monitor", 30, 35, 40)]
for row in data:
    ws.append(list(row) + [None])
for r in range(2, 6):
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})")
ws.append(["Promedio", "=AVERAGE(B2:B5)", "=AVERAGE(C2:C5)", "=AVERAGE(D2:D5)", ""])
ws.append(["Maximo", "=MAX(B2:B5)", "=MAX(C2:C5)", "=MAX(D2:D5)", ""])
style_header(ws, 1, 5)
auto_width(ws)
save(wb, "M04_Formulas_Funciones.xlsx")

# --- M5: Analisis de Datos ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ventas"
ws.append(["ID", "Vendedor", "Region", "Producto", "Ventas", "Mes"])
data = [
    (101, "Ana", "Norte", "Laptop", 15000, "Enero"),
    (102, "Luis", "Sur", "Mouse", 8000, "Enero"),
    (103, "Carlos", "Norte", "Teclado", 5000, "Febrero"),
    (104, "Marta", "Centro", "Monitor", 12000, "Febrero"),
    (105, "Pedro", "Sur", "Laptop", 18000, "Marzo"),
    (106, "Sofia", "Norte", "Mouse", 6500, "Marzo"),
]
for row in data:
    ws.append(row)
style_header(ws, 1, 6)
auto_width(ws)
save(wb, "M05_Analisis_Datos.xlsx")

# --- M6: Busqueda ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos"
ws.append(["ID", "Producto", "Precio", "Categoria"])
data = [
    (201, "Laptop Pro", 1500, "Electronica"),
    (202, "Mouse Pro", 45, "Accesorios"),
    (203, "Teclado Mecanico", 120, "Accesorios"),
    (204, "Monitor 27p", 400, "Electronica"),
]
for row in data:
    ws.append(row)
style_header(ws, 1, 4)
auto_width(ws)

# Hoja de busqueda
ws2 = wb.create_sheet("BuscarV")
ws2["A1"] = "ID a buscar"
ws2["B1"] = "Resultado"
ws2["A2"] = 202
ws2["B2"] = '=VLOOKUP(A2, Productos!A:D, 2, FALSE)'
ws2["A3"] = 204
ws2["B3"] = '=VLOOKUP(A3, Productos!A:D, 3, FALSE)'
style_header(ws2, 1, 2)
auto_width(ws2)
save(wb, "M06_Busqueda.xlsx")

# --- M7: Visualizacion ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ventas Mensuales"
ws.append(["Mes", "Ventas"])
data = [("Ene", 12000), ("Feb", 15000), ("Mar", 11000), ("Abr", 16500), ("May", 14000), ("Jun", 18500)]
for row in data:
    ws.append(row)
style_header(ws, 1, 2)
chart = BarChart()
chart.title = "Ventas por Mes"
chart.y_axis.title = "Ventas ($)"
chart.x_axis.title = "Mes"
data_ref = Reference(ws, min_col=2, min_row=1, max_row=7)
cats = Reference(ws, min_col=1, min_row=2, max_row=7)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, "D2")
auto_width(ws)
save(wb, "M07_Visualizacion.xlsx")

# --- M8: Avanzadas ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Prestamo"
ws["A1"] = "Monto"; ws["B1"] = 25000
ws["A2"] = "Tasa Anual"; ws["B2"] = 0.08
ws["A3"] = "Plazo (meses)"; ws["B3"] = 48
ws["A5"] = "Pago Mensual"; ws["B5"] = '=PMT(B2/12, B3, -B1)'
ws["A6"] = "Total Pagado"; ws["B6"] = '=B5*B3'
ws["A7"] = "Interes Total"; ws["B7"] = '=B6-B1'
hfmt = Font(bold=True, color="6B21A8")
for r in [1,2,3,5,6,7]:
    ws.cell(row=r, column=1).font = hfmt
auto_width(ws)
ws.column_dimensions["A"].width = 18
ws2 = wb.create_sheet("Escenarios")
ws2["A1"] = "Escenario"; ws2["B1"] = "Ventas Esperadas"
data = [("Pesimista", 80000), ("Realista", 120000), ("Optimista", 160000)]
for row in data:
    ws2.append(row)
style_header(ws2, 1, 2)
auto_width(ws2)
save(wb, "M08_Avanzadas.xlsx")

# --- M9: Proyecto Final ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Dashboard"
ws.append(["Producto", "Cantidad", "Precio Unitario", "Total"])
data = [
    ("Laptop", 15, 1200),
    ("Monitor", 10, 350),
    ("Teclado", 25, 80),
    ("Mouse", 30, 25),
    ("Webcam", 12, 90),
    ("Audifonos", 20, 60),
]
for p, c, pu in data:
    ws.append([p, c, pu, c*pu])
style_header(ws, 1, 4)
auto_width(ws)
save(wb, "M09_Proyecto_Final.xlsx")

print("\nTodos los archivos generados en:", OUT)
